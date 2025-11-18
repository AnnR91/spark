import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.utils import get_column_letter

def save_df_as_formatted_excel(df, output_file='output.xlsx'):
    """
    Save DataFrame to Excel with specific formatting matching the template.

    Parameters:
    -----------
    df : pd.DataFrame
        DataFrame with columns: 'text', 't1', 't2', 't3', 't4', 't5', 't6', 'text2', 'text3'
        Note: Column 'text3' will be populated based on 'text2' values
    output_file : str
        Output Excel file path
    """

    # Calculate column I (text3) based on column H (text2)
    # If text2 <= 2: small, <= 4: medium, > 4: large
    def categorize_value(val):
        if pd.isna(val):
            return ''
        if val <= 2:
            return 'small'
        elif val <= 4:
            return 'medium'
        else:
            return 'large'

    # Apply the categorization to text3 column if text2 exists
    if 'text2' in df.columns:
        if 'text3' not in df.columns:
            df['text3'] = df['text2'].apply(categorize_value)
        else:
            # Update text3 based on text2
            df['text3'] = df['text2'].apply(categorize_value)

    # Save DataFrame to Excel without headers (we'll add custom headers)
    df.to_excel(output_file, index=False, header=False, sheet_name='Arkusz1', startrow=3)

    # Load workbook to add formatting
    wb = load_workbook(output_file)
    ws = wb.active

    # Add header rows (rows 1-3)
    # Row 1: Main headers
    ws['A1'] = 'text'
    ws['B1'] = 'text1'
    ws['H1'] = 'text2'
    ws['I1'] = 'text3'

    # Row 2: Sub-headers
    ws['B2'] = 't1'
    ws['C2'] = 't2'
    ws['D2'] = 't3'
    ws['E2'] = 't4'
    ws['F2'] = 't5'
    ws['G2'] = 't6'

    # Row 3: Values
    ws['B3'] = 0.01
    ws['C3'] = 0.02
    ws['D3'] = 0.03
    ws['E3'] = 0.04
    ws['F3'] = 0.05
    ws['G3'] = 0.06

    # Apply conditional formatting for S/M/L values in columns B-G (starting from row 4)
    # Get the last row with data
    last_row = len(df) + 3  # +3 because we start from row 4 (3 header rows)

    # Define fills for S, M, L
    fill_s = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Light red
    fill_m = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Light yellow
    fill_l = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light green

    # Create conditional formatting rules for columns B through G
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        cell_range = f'{col}4:{col}{last_row}'

        # Rule for 'S' values
        rule_s = CellIsRule(operator='equal', formula=['"S"'], fill=fill_s)
        ws.conditional_formatting.add(cell_range, rule_s)

        # Rule for 'M' values
        rule_m = CellIsRule(operator='equal', formula=['"M"'], fill=fill_m)
        ws.conditional_formatting.add(cell_range, rule_m)

        # Rule for 'L' values
        rule_l = CellIsRule(operator='equal', formula=['"L"'], fill=fill_l)
        ws.conditional_formatting.add(cell_range, rule_l)

    # Apply color scale formatting to column I (text3)
    color_scale_range = f'I4:I{last_row}'
    color_scale_rule = ColorScaleRule(
        start_type='min',
        start_color='FFFFFFFF',  # White
        end_type='max',
        end_color='FF57BB8A'  # Green
    )
    ws.conditional_formatting.add(color_scale_range, color_scale_rule)

    # Adjust column widths for better readability
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    wb.save(output_file)
    print(f"Excel file saved successfully to: {output_file}")
    return output_file


# Example usage:
if __name__ == "__main__":
    # Create sample DataFrame
    sample_data = {
        'text': ['A1', 'A2', 'A3'],
        't1': ['S', 'M', 'L'],
        't2': ['S', 'M', 'L'],
        't3': ['S', 'M', 'L'],
        't4': ['S', 'M', 'L'],
        't5': ['S', 'M', 'L'],
        't6': ['S', 'M', 'L'],
        'text2': [1.0, 3.0, 5.0],
        # 'text3' will be automatically calculated based on 'text2'
    }

    df = pd.DataFrame(sample_data)

    # Save with formatting
    save_df_as_formatted_excel(df, 'output.xlsx')

    print("\nDataFrame preview:")
    print(df)
